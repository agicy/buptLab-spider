import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas
from scrapy import spiderloader
from scrapy.crawler import CrawlerProcess
from scrapy.utils.project import get_project_settings


def run_spider() -> None:
    settings = get_project_settings()
    settings.set("LOG_LEVEL", "INFO", priority="cmdline")

    process = CrawlerProcess(settings)
    spider_loader = spiderloader.SpiderLoader.from_settings(settings)

    print(spider_loader.list())

    crawlers = []
    for spider in spider_loader.list():
        print(f"Running spider {spider}")
        crawler = process.create_crawler(spider)
        crawlers.append(crawlers)
        process.crawl(crawler)

    process.start()


def preprocess_data_frame(data_frame: pandas.DataFrame) -> pandas.DataFrame:
    data_frame["ad_topic"] = data_frame["ad_topic"].str.strip()
    data_frame["post_date"] = pandas.to_datetime(data_frame["post_date"]).dt.strftime(
        r"%Y-%m-%d"
    )
    data_frame["visit_count"] = data_frame["visit_count"].fillna(0).astype(int)
    data_frame = data_frame.rename(
        columns={
            "ad_topic": "招聘主题",
            "post_date": "发布日期",
            "visit_count": "浏览次数",
        }
    )
    return data_frame


def preprocess(info_dict: dict[str, str]) -> dict[str, pandas.DataFrame]:
    result: dict[str, pandas.DataFrame] = {}
    for school, filename in info_dict.items():
        data_frame = pandas.read_csv(filename)
        data_frame = preprocess_data_frame(
            data_frame,
        )
        result[school] = data_frame

    return result


def fill_in_sheet(
    worksheet,
    data: pandas.DataFrame,
    header: list[str],
    header_style: str,
    column_styles: list[str],
    width_list: list[int],
) -> None:
    worksheet.append(header)
    for col, _ in enumerate(header, start=1):
        worksheet.cell(row=1, column=col).style = header_style
    for id, data in enumerate(
        dataframe_to_rows(data, index=False, header=False),
        start=1,
    ):
        values = [id] + list(data)
        worksheet.append(values)

        for col, _ in enumerate(values, start=1):
            worksheet.cell(row=id + 1, column=col).style = column_styles[col - 1]

    for col, width in enumerate(width_list, start=1):
        col_letter = worksheet.cell(row=1, column=col).column_letter
        worksheet.column_dimensions[col_letter].width = width


if __name__ == "__main__":
    run_spider()

    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [main] %(levelname)s %(message)s"
    )
    logger = logging.getLogger(__name__)

    logger.info("Preprocessing data...")
    preprocessed_data: dict[str, pandas.DataFrame] = preprocess(
        {
            "北邮": "data/bupt.csv",
            "西电": "data/xidian.csv",
        }
    )
    logger.info("Preprocessing finished.")

    logger.info("Generating Excel file...")

    header_style = NamedStyle(
        name="表头样式",
        font=Font(name="黑体", bold=True),
        alignment=Alignment(horizontal="left"),
        number_format="@",
    )
    text_style = NamedStyle(
        name="文本样式",
        font=Font(name="黑体"),
        alignment=Alignment(horizontal="left"),
        number_format="@",
    )
    date_style = NamedStyle(
        name="日期样式",
        font=Font(name="Times New Roman"),
        number_format="YYYY-MM-DD",
    )
    number_style = NamedStyle(
        name="数字样式",
        font=Font(name="Times New Roman"),
        number_format="0",
    )

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.add_named_style(header_style)
    workbook.add_named_style(text_style)
    workbook.add_named_style(date_style)
    workbook.add_named_style(number_style)

    header = ["序号", "招聘主题", "发布日期", "浏览次数"]
    width_list = [10, 40, 20, 10]

    k = 10
    fill_in_sheet(
        workbook.create_sheet("北邮"),
        preprocessed_data["北邮"],
        ["序号", "招聘主题", "发布日期", "浏览次数"],
        "表头样式",
        ["文本样式", "文本样式", "日期样式", "数字样式"],
        [10, 40, 20, 10],
    )

    fill_in_sheet(
        workbook.create_sheet("西电"),
        preprocessed_data["西电"],
        ["序号", "招聘主题", "发布日期", "浏览次数"],
        "表头样式",
        ["文本样式", "文本样式", "日期样式", "数字样式"],
        [10, 40, 20, 10],
    )

    fill_in_sheet(
        workbook.create_sheet(f"北邮招聘TOP{k}"),
        preprocessed_data["北邮"].nlargest(k, "浏览次数"),
        ["序号", "招聘主题", "发布日期", "浏览次数"],
        "表头样式",
        ["文本样式", "文本样式", "日期样式", "数字样式"],
        [10, 40, 20, 10],
    )

    fill_in_sheet(
        workbook.create_sheet(f"西电招聘TOP{k}"),
        preprocessed_data["西电"].nlargest(k, "浏览次数"),
        ["序号", "招聘主题", "发布日期", "浏览次数"],
        "表头样式",
        ["文本样式", "文本样式", "日期样式", "数字样式"],
        [10, 40, 20, 10],
    )

    common_jobs = pandas.merge(
        preprocessed_data["北邮"],
        preprocessed_data["西电"],
        on="招聘主题",
        suffixes=("_北邮", "_西电"),
    )
    common_jobs["总浏览次数"] = (
        common_jobs["浏览次数_北邮"] + common_jobs["浏览次数_西电"]
    )
    common_topk = common_jobs.nlargest(k, "总浏览次数")

    fill_in_sheet(
        workbook.create_sheet(f"两校招聘TOP{k}"),
        common_topk,
        [
            "序号",
            "招聘主题",
            "发布日期_北邮",
            "浏览次数_北邮",
            "发布日期_西电",
            "浏览次数_西电",
            "总浏览次数",
        ],
        "表头样式",
        [
            "文本样式",
            "文本样式",
            "日期样式",
            "数字样式",
            "日期样式",
            "数字样式",
            "数字样式",
        ],
        [10, 40, 20, 20, 20, 20, 10],
    )

    workbook.save("data/就业信息汇总.xlsx")

    logger.info("Excel file generated.")
